import openpyxl
from openpyxl import Workbook

new = Workbook()
ans = new.active

f = open("../output2.txt")
ans.cell(1,1).value = "平均到达时间间隔"
ans.cell(2,1).value = "平均服务时间"
ans.cell(3,1).value = "最大队列长度"
ans.cell(4,1).value = "服务器数目"
ans.cell(5,1).value = "结束时间"
ans.cell(6,1).value = "顾客数目"
ans.cell(7,1).value = "服务顾客数"
ans.cell(8,1).value = "队列中平均等待客户数"
ans.cell(9,1).value = "平均等待时间"
ans.cell(10,1).value = "平均逗留时间"
ans.cell(11,1).value = "实际平均服务时间"
ans.cell(12,1).value = "服务器利用率"
ans.cell(13,1).value = "总时间"

count_y = 2
line = f.readline()
while line:
    temp = line.split(' ')
    count_x = 1
    for str in temp:
        temp_str = str[0:6]
        ans.cell(count_x, count_y).value = temp_str
        count_x += 1
    line = f.readline()
    count_y += 1
f.close()

new.save(r'compare_mm1_2.xlsx')