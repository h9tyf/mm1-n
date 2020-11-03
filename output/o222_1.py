import openpyxl
from openpyxl import Workbook

new = Workbook()
ans = new.active

ans.cell(1,1).value = "time"
ans.cell(1,2).value = "0.1: 总人数"
ans.cell(1,3).value = "0.1: 已经服务人数"
ans.cell(1,4).value = "0.1: 排队人数"
ans.cell(1,5).value = "time2"
ans.cell(1,6).value = "0.2: 总人数"
ans.cell(1,7).value = "0.2: 已经服务人数"
ans.cell(1,8).value = "0.2: 排队人数"
ans.cell(1,9).value = "time3"
ans.cell(1,10).value = "0.3: 总人数"
ans.cell(1,11).value = "0.2: 已经服务人数"
ans.cell(1,12).value = "0.2: 排队人数"

f = open("../output_222_1.txt")
count_x = 2

line = f.readline()
while line:
    line.replace('\n', '')
    temp = line.split(' ')
    count_y = 1
    for str in temp:
        temp_str = str[0:6]
        ans.cell(count_x, count_y).value = temp_str
        count_y += 1
    line = f.readline()
    count_x += 1
f.close()

f = open("../output_222_2.txt")
count_x = 2
line = f.readline()
while line:
    line.replace('\n', '')
    temp = line.split(' ')
    count_y = 5
    for str in temp:
            temp_str = str[0:6]
            ans.cell(count_x, count_y).value = temp_str
            count_y += 1

    line = f.readline()
    count_x += 1
f.close()

f = open("../output_222_3.txt")
count_x = 2
line = f.readline()
while line:
    line.replace('\n', '')
    temp = line.split(' ')
    count_y = 9
    for str in temp:
            temp_str = str[0:6]
            ans.cell(count_x, count_y).value = temp_str
            count_y += 1
    line = f.readline()
    count_x += 1
f.close()

new.save(r'compare_mm1_222.xlsx')